"""
AI Cyber Assurance Agent — Streamlit web interface.

Run with:
    streamlit run app.py
"""

import io
import json
import os
import csv
import re
from pathlib import Path

import streamlit as st
from dotenv import load_dotenv

# ── Page config must be the very first Streamlit call ──────────────────────────
st.set_page_config(
    page_title="AI Cyber Assurance Agent",
    page_icon="🔐",
    layout="wide",
    initial_sidebar_state="expanded",
)

load_dotenv()

SPREADSHEET_STORAGE_DIR = Path(__file__).resolve().parent / "data"
PERSISTED_SPREADSHEET_BASE = SPREADSHEET_STORAGE_DIR / "uploaded_controls"

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
    :root {
        color-scheme: dark;
        color: #f8fafc;
        font-family: Inter, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }

    section.main {
        background: linear-gradient(180deg, #090e1d 0%, #111827 50%, #0b1226 100%);
    }

    .block-container {
        padding-top: 1rem;
        padding-bottom: 1rem;
    }

    .main-header {
        background: rgba(15, 23, 42, 0.95);
        border: 1px solid rgba(148, 163, 184, 0.18);
        box-shadow: 0 30px 60px rgba(15, 23, 42, 0.24);
        backdrop-filter: blur(20px);
        padding: 1.9rem 2.4rem;
        border-radius: 24px;
        margin-bottom: 1.5rem;
    }

    .main-header h1 {
        margin: 0;
        font-size: 2.6rem;
        letter-spacing: -0.04em;
        line-height: 1.05;
    }

    .main-header p {
        margin: 0.75rem 0 0 0;
        opacity: 0.82;
        font-size: 1.05rem;
        max-width: 760px;
        line-height: 1.75;
    }

    .hero-chip {
        display: inline-flex;
        align-items: center;
        margin-top: 1.15rem;
        padding: 0.62rem 1rem;
        border-radius: 999px;
        border: 1px solid rgba(148, 163, 184, 0.18);
        background: rgba(148, 163, 184, 0.09);
        color: #dbeafe;
        font-size: 0.95rem;
        font-weight: 700;
    }

    .sidebar-card {
        background: rgba(15, 23, 42, 0.97);
        border: 1px solid rgba(148, 163, 184, 0.18);
        border-radius: 22px;
        padding: 1.4rem 1.5rem;
        margin-bottom: 1rem;
        box-shadow: 0 16px 40px rgba(15, 23, 42, 0.22);
    }

    .sidebar-card h2 {
        margin: 0 0 0.85rem 0;
        color: #f8fafc;
        font-size: 1.15rem;
    }

    .sidebar-card p,
    .sidebar-card li {
        color: #cbd5e1;
        line-height: 1.75;
    }

    .sidebar-card ul {
        padding-left: 1.2rem;
        margin: 0.35rem 0 0 0;
    }

    .st-expander {
        background: rgba(15, 23, 42, 0.88);
        border: 1px solid rgba(148, 163, 184, 0.16);
        border-radius: 18px;
        padding: 0.85rem 1.05rem 1rem 1.05rem;
        margin-bottom: 1rem;
    }

    .st-expander header {
        font-weight: 700;
    }

    .stTextArea textarea,
    .stTextInput>div>div>input,
    .stTextArea>div>div>textarea {
        border-radius: 18px !important;
        border: 1px solid rgba(148, 163, 184, 0.18) !important;
        background: rgba(15, 23, 42, 0.96) !important;
        color: #e2e8f0 !important;
        padding: 1rem !important;
        box-shadow: inset 0 2px 10px rgba(15, 23, 42, 0.35) !important;
    }

    textarea:focus,
    input:focus {
        outline: 2px solid rgba(59, 130, 246, 0.34) !important;
        outline-offset: 0px !important;
    }

    .stButton>button {
        border-radius: 16px !important;
        padding: 1rem 1.5rem !important;
        font-size: 1rem !important;
        font-weight: 700 !important;
        background: linear-gradient(135deg, #2563eb 0%, #38bdf8 100%) !important;
        border: 1px solid rgba(96, 165, 250, 0.38) !important;
        box-shadow: 0 22px 40px rgba(59, 130, 246, 0.18) !important;
        color: #ffffff !important;
    }

    .stButton>button:hover {
        transform: translateY(-1px);
        box-shadow: 0 24px 42px rgba(59, 130, 246, 0.24) !important;
    }

    .stButton>button:active {
        transform: translateY(0);
    }

    .badge-met,
    .badge-partial,
    .badge-not-met {
        display: inline-flex;
        align-items: center;
        gap: 0.45rem;
        padding: 0.85rem 1.1rem;
        border-radius: 999px;
        font-weight: 700;
        font-size: 1rem;
        margin-bottom: 1.2rem;
    }

    .badge-met {
        color: #0f172a;
        background: #dcfce7;
    }

    .badge-partial {
        color: #92400e;
        background: #fef3c7;
    }

    .badge-not-met {
        color: #7f1d1d;
        background: #fee2e2;
    }

    .card-gap,
    .card-rec,
    .card-q {
        border-radius: 18px;
        padding: 1rem 1.15rem;
        margin-bottom: 1rem;
        box-shadow: 0 18px 38px rgba(15, 23, 42, 0.16);
        line-height: 1.75;
    }

    .card-gap {
        background: rgba(254, 243, 199, 0.14);
        border: 1px solid rgba(253, 224, 71, 0.18);
    }

    .card-rec {
        background: rgba(220, 252, 231, 0.16);
        border: 1px solid rgba(34, 197, 94, 0.18);
    }

    .card-q {
        background: rgba(237, 233, 254, 0.16);
        border: 1px solid rgba(139, 92, 246, 0.18);
    }

    .card-label {
        font-weight: 700;
        margin-right: 0.55rem;
    }

    .stTabs [role="tab"] {
        border-radius: 999px !important;
        background: rgba(148, 163, 184, 0.08) !important;
        margin-right: 0.35rem !important;
        padding: 0.75rem 1rem !important;
        min-width: 140px;
    }

    .stTabs [role="tab"][aria-selected="true"] {
        background: rgba(96, 165, 250, 0.22) !important;
        color: #ffffff !important;
        font-weight: 700 !important;
    }

    .footer-text {
        color: #94a3b8;
        text-align: center;
        margin-top: 1.5rem;
        padding: 0.75rem 0;
    }

    .stDivider>div>hr {
        border-top: 1px solid rgba(148, 163, 184, 0.16) !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <div class="main-header">
        <div style="display:flex; align-items:flex-start; justify-content:space-between; gap:1rem; flex-wrap:wrap;">
            <div style="max-width:720px;">
                <h1>🔐 AI Cyber Assurance Agent</h1>
                <p>Modern evidence-driven security control assessment powered by a structured AI audit workflow. Review controls, upload evidence, and generate defensible findings in one polished interface.</p>
                <span class="hero-chip">Auditor → Reviewer → Questioning</span>
            </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ── API key guard ──────────────────────────────────────────────────────────────
if not os.getenv("ANTHROPIC_API_KEY"):
    st.error(
        "⚠️  **ANTHROPIC_API_KEY not found.**  "
        "Please add it to your `.env` file and restart the app."
    )
    st.stop()

with st.sidebar:
    st.markdown(
        """
        <div class="sidebar-card">
            <h2>Quick Start</h2>
            <p>Load controls and evidence, complete the input form, then run the audit assessment.</p>
            <ul>
                <li>Upload spreadsheets to populate control rows.</li>
                <li>Upload multiple evidence files for combined input.</li>
                <li>Use the sample loader to test the workflow.</li>
            </ul>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div class="sidebar-card">
            <h2>Why this works</h2>
            <p>The structured AI pipeline gives you consistent, evidence-based output that is easier to review and defend.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

# ── Session state initialisation ──────────────────────────────────────────────
for key, default in {
    "form_control": "",
    "form_requirement": "",
    "form_evidence": "",
    "audit_result": None,
    "audit_error": None,
    "pipeline": None,
    "spreadsheet_rows": [],
    "spreadsheet_file_name": "",
    "sheet_row_index": 0,
    "uploaded_evidence_names": [],
}.items():
    if key not in st.session_state:
        st.session_state[key] = default

# Initialise pipeline once per session
if st.session_state.pipeline is None:
    try:
        from agents.pipeline import AuditPipeline

        st.session_state.pipeline = AuditPipeline()
    except Exception as exc:
        st.error(f"Failed to initialise audit pipeline: {exc}")
        st.stop()


# ── Sample loader helper ───────────────────────────────────────────────────────
def _load_sample() -> None:
    sample_path = Path(__file__).resolve().parent / "data" / "sample_input.json"
    try:
        sample = json.loads(sample_path.read_text(encoding="utf-8"))
        st.session_state.form_control = sample["control"]
        st.session_state.form_requirement = sample["requirement"]
        st.session_state.form_evidence = sample["evidence"]
    except Exception as exc:
        st.warning(f"Could not load sample input: {exc}")


def _parse_spreadsheet(file_bytes: bytes, filename: str) -> list[dict]:
    """Parse a spreadsheet file into a list of row dictionaries."""
    filename = filename.lower()
    if filename.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(text.splitlines())
        return [dict(row) for row in reader]

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records = []
    for row in rows[1:]:
        if row is None or not any(cell is not None for cell in row):
            continue
        record = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        records.append(record)
    return records


def _load_spreadsheet_bytes(file_bytes: bytes, filename: str) -> list[dict]:
    rows = _parse_spreadsheet(file_bytes, filename)
    normalized_rows = [_normalize_spreadsheet_row(record) for record in rows]
    return [
        row
        for row in normalized_rows
        if row["control"] or row["requirement"] or row["evidence"]
    ]


def _save_persisted_spreadsheet(file_bytes: bytes, filename: str) -> None:
    PERSISTED_SPREADSHEET_BASE.parent.mkdir(parents=True, exist_ok=True)
    target_path = PERSISTED_SPREADSHEET_BASE.with_suffix(Path(filename).suffix)
    target_path.write_bytes(file_bytes)

    for ext in [".csv", ".xlsx"]:
        candidate = PERSISTED_SPREADSHEET_BASE.with_suffix(ext)
        if candidate.exists() and candidate != target_path:
            candidate.unlink()


def _find_persisted_spreadsheet() -> Path | None:
    for ext in [".xlsx", ".csv"]:
        path = PERSISTED_SPREADSHEET_BASE.with_suffix(ext)
        if path.exists():
            return path
    return None


def _load_persisted_spreadsheet() -> tuple[list[dict], str] | None:
    persisted_path = _find_persisted_spreadsheet()
    if not persisted_path:
        return None
    try:
        rows = _load_spreadsheet_bytes(persisted_path.read_bytes(), persisted_path.name)
        return rows, persisted_path.name
    except Exception:
        return None


# Load a persisted spreadsheet from disk if one exists and no sheet rows are currently loaded.
if not st.session_state.spreadsheet_rows:
    persisted = _load_persisted_spreadsheet()
    if persisted:
        st.session_state.spreadsheet_rows, st.session_state.spreadsheet_file_name = persisted


def _normalize_spreadsheet_row(record: dict) -> dict:
    normalized = {}
    for key, value in record.items():
        name = str(key or "").strip().lower()
        text = str(value).strip() if value is not None else ""
        normalized[name] = text

    control = ""
    requirement = ""
    evidence = ""

    def normalize_key(key: str) -> str:
        return "".join(ch for ch in key.lower() if ch.isalnum())

    for key, text in normalized.items():
        if not text:
            continue
        short_key = normalize_key(key)

        if short_key in {
            "control",
            "controlname",
            "controldescription",
            "c",
            "cdescription",
            "controllabel",
            "controltitle",
        } or "control" in short_key and "evidence" not in short_key:
            control = text
            continue

        if short_key in {
            "detailedadequacyrequirements",
            "adequacyrequirements",
            "adequacyrequirement",
            "detailedrequirement",
            "requirements",
            "requirement",
            "requirementdetails",
            "dx",
            "detailedadequacy",
        } or "requirement" in short_key or "adequacy" in short_key:
            requirement = text
            continue

        if short_key in {
            "evidencerequirements",
            "evidencerequirement",
            "evidence",
            "ex",
            "evidenceex",
            "evidencelabel",
        } or "evidence" in short_key:
            evidence = text
            continue

        if "description" in short_key and not requirement:
            requirement = text

    return {
        "control": control,
        "requirement": requirement,
        "evidence": evidence,
        "raw": record,
    }


def _load_spreadsheet(file) -> list[dict]:
    try:
        rows = _parse_spreadsheet(file.read(), file.name)
    except Exception as exc:
        st.warning(f"Could not parse spreadsheet: {exc}")
        return []

    normalized_rows = [_normalize_spreadsheet_row(record) for record in rows]
    return [
        row
        for row in normalized_rows
        if row["control"] or row["requirement"] or row["evidence"]
    ]


def _load_evidence_file(file) -> str | None:
    try:
        filename = file.name.lower()
        raw = file.read()

        if filename.endswith(".csv"):
            text = raw.decode("utf-8-sig") if isinstance(raw, bytes) else str(raw)
            reader = csv.reader(text.splitlines())
            rows = [", ".join(row) for row in reader]
            return "\n".join(rows)

        if filename.endswith(".json"):
            text = raw.decode("utf-8-sig") if isinstance(raw, bytes) else str(raw)
            try:
                data = json.loads(text)
                return json.dumps(data, indent=2, ensure_ascii=False)
            except json.JSONDecodeError:
                return text

        if filename.endswith(".pdf"):
            try:
                from pypdf import PdfReader
            except ImportError as exc:
                raise RuntimeError(
                    "PDF support requires the pypdf package. "
                    "Please install it in the virtual environment."
                ) from exc

            reader = PdfReader(io.BytesIO(raw) if isinstance(raw, bytes) else io.BytesIO(str(raw).encode("utf-8")))
            pages = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages).strip()

        if filename.endswith((".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif")):
            try:
                from PIL import Image
            except ImportError as exc:
                raise RuntimeError(
                    "Image support requires Pillow. "
                    "Please install Pillow in the virtual environment."
                ) from exc

            try:
                import pytesseract
            except ImportError as exc:
                raise RuntimeError(
                    "Image OCR requires pytesseract. "
                    "Please install pytesseract in the virtual environment."
                ) from exc

            image = Image.open(io.BytesIO(raw) if isinstance(raw, bytes) else io.BytesIO(str(raw).encode("utf-8")))
            try:
                text = pytesseract.image_to_string(image)
            except pytesseract.TesseractNotFoundError as exc:
                raise RuntimeError(
                    "Tesseract OCR executable not found. "
                    "Install Tesseract on your system and ensure it is available in PATH."
                ) from exc
            return text.strip()

        if isinstance(raw, bytes):
            try:
                return raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                return raw.decode("latin-1")

        return str(raw)
    except Exception as exc:
        st.warning(f"Could not load evidence file: {exc}")
        return None


def _format_markdown_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(f"- {item}" for item in value if item)
    if isinstance(value, dict):
        return "\n".join(f"- **{k}:** {v}" for k, v in value.items() if v)

    text = str(value).strip()
    if not text:
        return ""

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if any(re.match(r"^[-*•] ", line) or re.match(r"^\d+\. ", line) for line in lines):
        return "\n".join(lines)

    return text


def _load_evidence_files(files) -> str:
    if not files:
        return ""

    if not isinstance(files, list):
        files = [files]

    contents = []
    unsupported = []

    for file in files:
        filename = file.name.lower()
        if not any(
            filename.endswith(ext)
            for ext in [".txt", ".md", ".csv", ".json", ".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif"]
        ):
            unsupported.append(file.name)
            continue

        text = _load_evidence_file(file)
        if text is not None:
            contents.append(f"--- {file.name} ---\n{text}")

    if unsupported:
        st.warning(
            "Some uploaded files were skipped because their format is not supported: "
            + ", ".join(unsupported)
        )

    return "\n\n".join(contents)


# Sample loader (before inputs)
with st.expander("📥 Load from spreadsheet"):
    st.write(
        "Upload an Excel or CSV file with columns such as Control, Requirement/Description, and Evidence."
    )
    uploaded_file = st.file_uploader(
        "Upload controls spreadsheet",
        type=["xlsx", "csv"],
        help="Use your offline sheet to pre-populate the audit input fields.",
        key="spreadsheet_file",
    )

    if uploaded_file is not None:
        file_bytes = uploaded_file.read()
        if (
            uploaded_file.name != st.session_state.spreadsheet_file_name
            or not st.session_state.spreadsheet_rows
        ):
            st.session_state.spreadsheet_rows = _load_spreadsheet_bytes(
                file_bytes, uploaded_file.name
            )
            st.session_state.spreadsheet_file_name = uploaded_file.name
            st.session_state.sheet_row_index = 0
            _save_persisted_spreadsheet(file_bytes, uploaded_file.name)

    sheet_rows = st.session_state.spreadsheet_rows
    if sheet_rows:
        st.info(
            f"Loaded {len(sheet_rows)} rows from {st.session_state.spreadsheet_file_name}."
        )
        options = [
            f"{i + 1}. {row['control'][:80]} / {row['requirement'][:80]}"
            for i, row in enumerate(sheet_rows)
        ]
        selected_index = st.selectbox(
            "Select row to load",
            range(len(sheet_rows)),
            format_func=lambda i: options[i],
            key="sheet_row_index",
        )
        if st.button("Load selected row", use_container_width=False):
            selected = sheet_rows[selected_index]
            st.session_state.form_control = selected["control"]
            st.session_state.form_requirement = selected["requirement"]
            st.session_state.form_evidence = selected["evidence"]
            st.rerun()
        if st.session_state.spreadsheet_file_name:
            st.caption(f"Current spreadsheet: {st.session_state.spreadsheet_file_name}")
    elif uploaded_file is not None:
        st.warning(
            "No usable rows found in the uploaded spreadsheet. Make sure the file includes Control, Requirement/Description, and Evidence columns."
        )

with st.expander("📂 Upload evidence files"):
    st.write(
        "Upload one or more evidence files and have them populate the Evidence field for audit use."
    )
    st.markdown(
        "_Supports text, JSON, CSV, PDF, and image screenshots. Image files are OCR scanned automatically._"
    )
    evidence_files = st.file_uploader(
        "Upload evidence files",
        type=None,
        accept_multiple_files=True,
        help="Upload evidence files to populate the Evidence input. Multiple files are supported.",
        key="evidence_files",
    )
    if evidence_files:
        if st.button("Load evidence files", use_container_width=False):
            evidence_text = _load_evidence_files(evidence_files)
            if evidence_text:
                if st.session_state.form_evidence:
                    st.session_state.form_evidence += "\n\n" + evidence_text
                else:
                    st.session_state.form_evidence = evidence_text
                names = [file.name for file in evidence_files]
                st.session_state.uploaded_evidence_names = names
                st.success(f"Evidence loaded from: {', '.join(names)}")
    if st.session_state.uploaded_evidence_names:
        st.caption(
            f"Loaded evidence files: {', '.join(st.session_state.uploaded_evidence_names)}"
        )

with st.expander("📎 Load sample input"):
    st.write("Click below to populate the form with a pre-built MFA control example.")
    if st.button("Load Sample", use_container_width=False):
        _load_sample()
        st.rerun()

# ── Input Section ──────────────────────────────────────────────────────────────
st.subheader("Audit Input")

top_row = st.columns([1, 1])

with top_row[0]:
    st.text_area(
        "Control Description [C]",
        key="form_control",
        value=st.session_state.form_control,
        height=130,
        placeholder=(
            "Describe the security control being assessed.\n\n"
            "e.g., Access Control — Multi-Factor Authentication: All privileged user "
            "accounts must be protected with MFA."
        ),
        help="The security control that is under assessment.",
    )
    st.text_area(
        "Requirement [Dx]",
        key="form_requirement",
        value=st.session_state.form_requirement,
        height=130,
        placeholder=(
            "State the specific requirement that must be evidenced.\n\n"
            "e.g., All privileged users must authenticate using MFA before accessing "
            "any critical system."
        ),
        help="The specific requirement that the evidence must demonstrate is met.",
    )

with top_row[1]:
    st.text_area(
        "Evidence [Ex]",
        key="form_evidence",
        value=st.session_state.form_evidence,
        height=280,
        placeholder=(
            "Describe or paste the evidence provided.\n\n"
            "e.g., Screenshot of Azure AD Conditional Access policy showing MFA "
            "enforcement for Global Administrators. Policy last modified 3 months ago."
        ),
        help="The evidence provided to demonstrate compliance with the requirement.",
    )

# ── Run button ─────────────────────────────────────────────────────────────────
st.divider()

col_btn, col_info = st.columns([2, 5])
with col_btn:
    run_clicked = st.button(
        "🔍 Run Audit Assessment",
        type="primary",
        use_container_width=True,
    )
with col_info:
    st.caption(
        "The pipeline runs three AI agents in sequence (Auditor → Reviewer → Questioning). "
        "This typically takes 20–40 seconds."
    )

if run_clicked:
    ctrl = st.session_state.form_control.strip()
    req = st.session_state.form_requirement.strip()
    ev = st.session_state.form_evidence.strip()

    if not ctrl or not req or not ev:
        st.error("Please fill in **all three fields** before running the assessment.")
    else:
        with st.spinner("Running multi-agent audit assessment…"):
            try:
                result = st.session_state.pipeline.run(
                    {"control": ctrl, "requirement": req, "evidence": ev}
                )
                st.session_state.audit_result = result
                st.session_state.audit_error = None
            except ValueError as exc:
                st.session_state.audit_error = str(exc)
                st.session_state.audit_result = None
            except Exception as exc:
                st.session_state.audit_error = (
                    f"An unexpected error occurred: {exc}"
                )
                st.session_state.audit_result = None

# ── Error display ──────────────────────────────────────────────────────────────
if st.session_state.audit_error:
    st.error(f"**Error:** {st.session_state.audit_error}")

# ── Results Section ────────────────────────────────────────────────────────────
if st.session_state.audit_result:
    result = st.session_state.audit_result

    st.divider()
    st.subheader("Audit Results")

    # Assessment badge
    assessment = result.get("requirement_assessment", "Unknown")
    badge_class = {
        "Met": "badge-met",
        "Partially Met": "badge-partial",
        "Not Met": "badge-not-met",
    }.get(assessment, "badge-partial")
    icon = {"Met": "✅", "Partially Met": "⚠️", "Not Met": "❌"}.get(assessment, "❓")

    st.markdown(
        f'<p class="{badge_class}">{icon}&nbsp; Requirement Assessment: {assessment}</p>',
        unsafe_allow_html=True,
    )
    st.write("")

    # ── Tabs ──────────────────────────────────────────────────────────────────
    tab_summary, tab_evidence, tab_gaps, tab_questions, tab_json = st.tabs(
        [
            "📋 Summary",
            "🔬 Evidence Quality",
            "⚠️ Gaps & Recommendations",
            "❓ Follow-up Questions",
            "📄 Raw JSON",
        ]
    )

    # --- Summary ---
    with tab_summary:
        st.markdown("**Control Objective**")
        st.markdown(_format_markdown_text(result.get("control_objective", "N/A")))

        st.markdown("**Audit Opinion**")
        st.markdown(_format_markdown_text(result.get("audit_opinion", "N/A")))

        summary_points = result.get("summary_points") or result.get("key_findings") or []
        if summary_points:
            st.markdown("**Key Findings**")
            st.markdown(_format_markdown_text(summary_points))

    # --- Evidence Quality ---
    with tab_evidence:
        eq = result.get("evidence_quality", {})
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**Completeness**")
            st.write(eq.get("completeness", "N/A"))
        with c2:
            st.markdown("**Relevance**")
            st.write(eq.get("relevance", "N/A"))
        with c3:
            st.markdown("**Reliability**")
            st.write(eq.get("reliability", "N/A"))

    # --- Gaps & Recommendations ---
    with tab_gaps:
        g_col, r_col = st.columns(2)

        with g_col:
            gaps = result.get("gaps_identified", [])
            st.markdown(f"**Gaps Identified ({len(gaps)})**")
            if gaps:
                for gap in gaps:
                    st.markdown(
                        f'<div class="card-gap"><span class="card-label">⚠️</span>{gap}</div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.success("No significant gaps identified.")

        with r_col:
            recs = result.get("recommendations", [])
            st.markdown(f"**Recommendations ({len(recs)})**")
            if recs:
                for rec in recs:
                    st.markdown(
                        f'<div class="card-rec"><span class="card-label">💡</span>{rec}</div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.info("No specific recommendations.")

    # --- Follow-up Questions ---
    with tab_questions:
        questions = result.get("follow_up_questions", [])
        if questions:
            st.markdown(f"**{len(questions)} follow-up question(s) generated:**")
            for i, q in enumerate(questions, 1):
                st.markdown(
                    f'<div class="card-q"><span class="card-label">Q{i}.</span>{q}</div>',
                    unsafe_allow_html=True,
                )
        else:
            st.success(
                "No follow-up questions required. "
                "The evidence is sufficient or the assessment is complete."
            )

    # --- Raw JSON ---
    with tab_json:
        json_str = json.dumps(result, indent=2, ensure_ascii=False)
        st.code(json_str, language="json")
        st.download_button(
            label="⬇️ Download JSON Report",
            data=json_str,
            file_name="audit_report.json",
            mime="application/json",
            use_container_width=False,
        )

# ── Footer ─────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    """
    <div class="footer-text">
        AI Cyber Assurance Agent · Powered by Anthropic Claude ·  All assessments are evidence-based and should be reviewed by a qualified auditor.
    </div>
    """,
    unsafe_allow_html=True,
)
